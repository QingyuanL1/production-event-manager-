import pandas as pd
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter import scrolledtext
import numpy as np

class DataLoader:
    def __init__(self, root):
        self.root = root
        self.root.title("生产排班系统 - 数据加载与预览")
        self.root.geometry("1200x800")
        
        # 设置表格样式
        self.setup_styles()
        
        self.data_frames = {}
        self.current_df = None
        self.current_file = None
        
        self.create_widgets()
        
    def setup_styles(self):
        # 创建表格样式
        style = ttk.Style()
        
        # 配置Treeview样式，添加边框
        style.configure("Treeview", 
                        background="#FFFFFF",
                        foreground="black",
                        rowheight=25,
                        fieldbackground="#FFFFFF",
                        borderwidth=1)
        
        # 配置Treeview.Heading样式（表头）
        style.configure("Treeview.Heading",
                        background="#E8E8E8",
                        foreground="black",
                        relief="raised",
                        borderwidth=1)
        
        # 配置单元格边框
        style.layout("Treeview", [
            ('Treeview.treearea', {'sticky': 'nswe', 'border': 1})
        ])
        
        # 配置选中行的样式
        style.map("Treeview",
                  background=[('selected', '#CCCCFF')])
        
    def create_widgets(self):
        # 创建顶部框架
        top_frame = ttk.Frame(self.root)
        top_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # 创建加载按钮
        ttk.Button(top_frame, text="加载数据表", command=self.load_file).pack(side=tk.LEFT, padx=5)
        
        # 创建文件选择下拉框
        self.file_var = tk.StringVar()
        self.file_combo = ttk.Combobox(top_frame, textvariable=self.file_var, state="readonly", width=30)
        self.file_combo.pack(side=tk.LEFT, padx=5)
        self.file_combo.bind("<<ComboboxSelected>>", self.on_file_selected)
        
        # 创建自动加载按钮
        ttk.Button(top_frame, text="自动加载所有数据表", command=self.load_all_files).pack(side=tk.LEFT, padx=5)
        
        # 创建工作表选择下拉框
        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(top_frame, textvariable=self.sheet_var, state="readonly", width=20)
        self.sheet_combo.pack(side=tk.LEFT, padx=5)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_selected)
        
        # 创建清理数据按钮
        ttk.Button(top_frame, text="清理数据", command=self.clean_data).pack(side=tk.LEFT, padx=5)
        
        # 创建导出按钮
        ttk.Button(top_frame, text="导出数据", command=self.export_data).pack(side=tk.LEFT, padx=5)
        
        # 创建信息标签
        self.info_label = ttk.Label(top_frame, text="")
        self.info_label.pack(side=tk.LEFT, padx=5)
        
        # 创建表格框架
        self.table_frame = ttk.Frame(self.root)
        self.table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 创建状态栏
        self.status_bar = ttk.Label(self.root, text="就绪", relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 默认加载数据表目录中的所有文件
        self.load_all_files()
        
    def load_file(self):
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls")]
        )
        
        if file_path:
            self.load_excel_file(file_path)
            
    def load_excel_file(self, file_path):
        try:
            self.status_bar.config(text=f"正在加载 {os.path.basename(file_path)}...")
            self.root.update()
            
            # 获取所有工作表
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            
            # 存储文件信息
            file_name = os.path.basename(file_path)
            self.data_frames[file_name] = {
                'path': file_path,
                'sheets': {},
                'sheet_names': sheet_names
            }
            
            # 加载第一个工作表
            if sheet_names:
                # 处理合并单元格
                df = self.read_excel_with_merged_cells(file_path, sheet_names[0])
                self.data_frames[file_name]['sheets'][sheet_names[0]] = df
                
                # 更新文件下拉框
                self.update_file_combo()
                
                # 选择当前加载的文件
                self.file_var.set(file_name)
                
                # 更新工作表下拉框
                self.update_sheet_combo(file_name)
                
                # 选择第一个工作表
                self.sheet_var.set(sheet_names[0])
                
                # 显示数据
                self.current_df = df
                self.current_file = file_name
                self.display_data(df)
                
                self.status_bar.config(text=f"已加载 {file_name} - {sheet_names[0]}")
                self.info_label.config(text=f"行数: {len(df)}, 列数: {len(df.columns)}")
            else:
                messagebox.showwarning("警告", f"文件 {file_name} 中没有工作表")
                self.status_bar.config(text="就绪")
        except Exception as e:
            messagebox.showerror("错误", f"加载文件时出错: {str(e)}")
            self.status_bar.config(text="加载失败")
            
    def read_excel_with_merged_cells(self, file_path, sheet_name):
        # 使用openpyxl引擎读取Excel
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
        
        try:
            # 获取工作簿和工作表
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            ws = wb[sheet_name]
            
            # 处理合并单元格
            for merged_range in ws.merged_cells.ranges:
                # 获取合并单元格的值
                top_left_cell_value = ws.cell(merged_range.min_row, merged_range.min_col).value
                
                # 将值填充到合并范围内的所有单元格
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        # 转换为pandas的行列索引
                        df_row = row - 1  # openpyxl是1-indexed，pandas是0-indexed
                        df_col = ws.cell(1, col).value  # 使用列标题
                        
                        # 确保列名存在且类型兼容
                        if df_col in df.columns and df_row < len(df):
                            try:
                                df.at[df_row, df_col] = top_left_cell_value
                            except:
                                # 如果类型不兼容，先转换为字符串
                                df.at[df_row, df_col] = str(top_left_cell_value) if top_left_cell_value is not None else None
        except Exception as e:
            print(f"处理合并单元格时出错: {e}")
        
        return df
    
    def load_all_files(self):
        # 默认数据表目录
        data_dir = os.path.join(os.getcwd(), "数据表")
        
        if os.path.exists(data_dir) and os.path.isdir(data_dir):
            excel_files = [f for f in os.listdir(data_dir) if f.endswith(('.xlsx', '.xls'))]
            
            if excel_files:
                for file_name in excel_files:
                    file_path = os.path.join(data_dir, file_name)
                    self.load_excel_file(file_path)
            else:
                messagebox.showinfo("信息", "数据表目录中没有Excel文件")
        else:
            messagebox.showwarning("警告", "找不到数据表目录")
            
    def update_file_combo(self):
        file_names = list(self.data_frames.keys())
        self.file_combo['values'] = file_names
        
    def update_sheet_combo(self, file_name):
        if file_name in self.data_frames:
            sheet_names = self.data_frames[file_name]['sheet_names']
            self.sheet_combo['values'] = sheet_names
        
    def on_file_selected(self, event):
        file_name = self.file_var.get()
        if file_name in self.data_frames:
            self.current_file = file_name
            self.update_sheet_combo(file_name)
            
            # 选择第一个工作表
            sheet_names = self.data_frames[file_name]['sheet_names']
            if sheet_names:
                self.sheet_var.set(sheet_names[0])
                
                # 如果工作表已经加载，则显示数据
                if sheet_names[0] in self.data_frames[file_name]['sheets']:
                    df = self.data_frames[file_name]['sheets'][sheet_names[0]]
                    self.current_df = df
                    self.display_data(df)
                    self.info_label.config(text=f"行数: {len(df)}, 列数: {len(df.columns)}")
                else:
                    # 加载工作表
                    self.load_sheet(file_name, sheet_names[0])
        
    def on_sheet_selected(self, event):
        file_name = self.file_var.get()
        sheet_name = self.sheet_var.get()
        
        if file_name in self.data_frames and sheet_name:
            # 检查工作表是否已加载
            if sheet_name in self.data_frames[file_name]['sheets']:
                df = self.data_frames[file_name]['sheets'][sheet_name]
                self.current_df = df
                self.display_data(df)
                self.info_label.config(text=f"行数: {len(df)}, 列数: {len(df.columns)}")
                self.status_bar.config(text=f"已加载 {file_name} - {sheet_name}")
            else:
                # 加载工作表
                self.load_sheet(file_name, sheet_name)
    
    def load_sheet(self, file_name, sheet_name):
        try:
            self.status_bar.config(text=f"正在加载 {file_name} - {sheet_name}...")
            self.root.update()
            
            file_path = self.data_frames[file_name]['path']
            df = self.read_excel_with_merged_cells(file_path, sheet_name)
            self.data_frames[file_name]['sheets'][sheet_name] = df
            
            self.current_df = df
            self.display_data(df)
            
            self.status_bar.config(text=f"已加载 {file_name} - {sheet_name}")
            self.info_label.config(text=f"行数: {len(df)}, 列数: {len(df.columns)}")
        except Exception as e:
            messagebox.showerror("错误", f"加载工作表时出错: {str(e)}")
            self.status_bar.config(text="加载失败")
    
    def clean_data(self):
        if self.current_df is None:
            messagebox.showinfo("信息", "请先加载数据")
            return
        
        try:
            df = self.current_df.copy()
            
            # 1. 向前填充NaN值（处理合并单元格）
            df = df.fillna(method='ffill')
            
            # 2. 删除全为NaN的行
            df = df.dropna(how='all')
            
            # 3. 删除全为NaN的列
            df = df.dropna(axis=1, how='all')
            
            # 4. 替换剩余的NaN为空字符串
            df = df.fillna('')
            
            # 更新当前数据框
            self.current_df = df
            
            # 更新存储的数据框
            file_name = self.current_file
            sheet_name = self.sheet_var.get()
            if file_name and sheet_name:
                self.data_frames[file_name]['sheets'][sheet_name] = df
            
            # 显示清理后的数据
            self.display_data(df)
            
            self.status_bar.config(text=f"已清理数据 - {file_name} - {sheet_name}")
            messagebox.showinfo("成功", "数据清理完成")
        except Exception as e:
            messagebox.showerror("错误", f"清理数据时出错: {str(e)}")
    
    def export_data(self):
        if self.current_df is None:
            messagebox.showinfo("信息", "请先加载数据")
            return
        
        try:
            file_path = filedialog.asksaveasfilename(
                title="保存Excel文件",
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx")]
            )
            
            if file_path:
                self.current_df.to_excel(file_path, index=False)
                self.status_bar.config(text=f"数据已导出到 {file_path}")
                messagebox.showinfo("成功", f"数据已成功导出到\n{file_path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出数据时出错: {str(e)}")
    
    def display_data(self, df):
        # 清除表格框架中的所有小部件
        for widget in self.table_frame.winfo_children():
            widget.destroy()
        
        # 创建一个框架，包含Treeview和滚动条
        frame = ttk.Frame(self.table_frame)
        frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建垂直滚动条
        vsb = ttk.Scrollbar(frame, orient="vertical")
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建水平滚动条
        hsb = ttk.Scrollbar(frame, orient="horizontal")
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 创建Treeview，添加网格线
        tree = ttk.Treeview(frame, yscrollcommand=vsb.set, xscrollcommand=hsb.set, style="Treeview")
        tree.pack(fill=tk.BOTH, expand=True)
        
        # 配置滚动条
        vsb.config(command=tree.yview)
        hsb.config(command=tree.xview)
        
        # 定义列
        columns = list(df.columns)
        tree["columns"] = columns
        tree["show"] = "headings"
        
        # 设置列标题和宽度
        for col in columns:
            tree.heading(col, text=str(col))
            # 根据内容设置列宽
            max_width = max(
                len(str(col)),
                df[col].astype(str).str.len().max() if len(df) > 0 else 0
            )
            width = min(max_width * 10, 300)  # 限制最大宽度
            tree.column(col, width=width, minwidth=50)
        
        # 添加数据，处理NaN值
        for i, row in df.iterrows():
            values = []
            for col in columns:
                val = row[col]
                if pd.isna(val):
                    values.append("")
                else:
                    values.append(str(val))
            
            # 为偶数行和奇数行设置不同的标签，以便应用不同的样式
            tag = "even" if i % 2 == 0 else "odd"
            tree.insert("", "end", values=values, tags=(tag,))
        
        # 设置行的交替颜色
        tree.tag_configure("even", background="#F0F0F0")
        tree.tag_configure("odd", background="#FFFFFF")
        
        # 创建数据预览文本框
        preview_frame = ttk.LabelFrame(self.table_frame, text="数据预览")
        preview_frame.pack(fill=tk.X, pady=10)
        
        # 创建文本框显示数据摘要
        text = scrolledtext.ScrolledText(preview_frame, height=10)
        text.pack(fill=tk.X)
        
        # 显示数据摘要，处理数值列
        try:
            numeric_df = df.select_dtypes(include=[np.number])
            if not numeric_df.empty:
                text.insert(tk.END, f"数值型数据摘要:\n\n{numeric_df.describe().to_string()}\n\n")
            else:
                text.insert(tk.END, "没有数值型数据列\n\n")
        except:
            text.insert(tk.END, "无法生成数值型数据摘要\n\n")
        
        # 显示数据类型
        text.insert(tk.END, f"数据类型:\n\n{df.dtypes.to_string()}\n\n")
        
        # 显示缺失值统计
        missing_data = df.isna().sum()
        missing_percent = (df.isna().sum() / len(df) * 100).round(2)
        missing_stats = pd.DataFrame({
            '缺失值数量': missing_data,
            '缺失值百分比(%)': missing_percent
        })
        text.insert(tk.END, f"缺失值统计:\n\n{missing_stats.to_string()}")
        
        text.config(state=tk.DISABLED)

if __name__ == "__main__":
    root = tk.Tk()
    app = DataLoader(root)
    root.mainloop() 